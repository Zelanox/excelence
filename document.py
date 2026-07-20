import pandas as pd

from typing import Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import network.protocol as protocol

class Document:

    def __init__(self, storage, network, online=False):

        # ==================================================
        # Managers
        # ==================================================

        self.storage = storage

        self.network = network

        self.online = online

        # ==================================================
        # Workbook
        # ==================================================

        self.workbook: Optional[Workbook] = None

        self.sheet: Optional[Worksheet] = None

        self.sheet_name = None

        self.sheet_names = []

        self.active_sheet_index = 0

        # ==================================================
        # Data
        # ==================================================

        self.df = pd.DataFrame()
        
        self.filtered_df = pd.DataFrame()
        
        self.auto_filter = None

        # ==================================================
        # Sheet Metadata
        # ==================================================

        self.merged_cells = []

        self.column_widths = {}

        self.row_heights = {}

        self.hidden_rows = set()

        self.hidden_columns = set()

        self.freeze_panes = None

        # ==================================================
        # Document State
        # ==================================================

        self.filename = None

        self.document_name = None

        self.modified = False

        self.loaded = False

        # ==================================================
        # Search / Sort
        # ==================================================

        self.search_text = ""

        self.sort_rules = []

        self.product_column = None

        # ==================================================
        # Collaboration
        # ==================================================

        self.pending_operations = []

        self.operation_history = []

        self.last_server_version = 0

        # ==================================================
        # Save State
        # ==================================================

        self.save_requested = False

        self.save_in_progress = False

    # ==========================================================
    # Open
    # ==========================================================

    def open(

        self,

        document_name

    ):

        self.document_name = document_name

        self.modified = False

        self.pending_operations.clear()

        self.operation_history.clear()

        if self.online:

            try:

                if not self.network.connect():

                    raise ConnectionError(
                        "Unable to connect to server."
                    )

                reply = self.network.begin_edit(
                    document_name
                )

                if reply["status"] != protocol.OK:

                    raise RuntimeError(

                        reply.get(

                            "message",

                            "Unable to open document."

                        )

                    )

                self.last_server_version = reply["version"]

                self.network.download_if_needed(

                    document_name,

                    self.last_server_version

                )

            except Exception as error:

                print("Network error:", error)

                print("Switching to offline mode.")

                self.online = False

                self.network.disconnect()

        self.load_local(document_name)

        self.loaded = True

        return True

    # ==========================================================
    # Create
    # ==========================================================

    def create(

        self,

        document_name

    ):

        self.document_name = document_name

        self.filename = document_name

        self.modified = False

        self.pending_operations.clear()

        self.operation_history.clear()

        # ------------------------------------------
        # Create Workbook
        # ------------------------------------------

        self.workbook = Workbook()

        self.sheet = self.workbook.active

        self.sheet.title = "Sheet1"

        # ------------------------------------------
        # Build DataFrame
        # ------------------------------------------

        self.df = pd.DataFrame()

        self.filtered_df = self.df

        # ------------------------------------------
        # Metadata
        # ------------------------------------------

        self.merged_cells.clear()

        self.column_widths.clear()

        self.row_heights.clear()

        self.hidden_rows.clear()

        self.hidden_columns.clear()

        self.freeze_panes = None

        # ------------------------------------------
        # Save locally
        # ------------------------------------------

        self.save_local()

        # ------------------------------------------
        # Upload to server (future)
        # ------------------------------------------

        if self.online:

            try:

                self.network.create_document(

                    document_name

                )

            except Exception as error:

                print(

                    "Unable to create server document:",

                    error

                )

        self.loaded = True

        return True

    def save_local(self):

        pass

    # ==========================================================
    # Load Local
    # ==========================================================

    def load_local(

        self,

        filename

    ):

        self.filename = filename

        # ------------------------------------------
        # Load workbook
        # ------------------------------------------

        self.workbook = self.storage.open_workbook(
            filename
        )

        # ------------------------------------------
        # Select active sheet
        # ------------------------------------------

        self.sheet = self.workbook.active

        self.sheet_name = self.sheet.title

        self.active_sheet_index = self.workbook.index(
            self.sheet
        )

        # ------------------------------------------
        # Load sheet metadata
        # ------------------------------------------

        self.load_metadata()

        # ------------------------------------------
        # Build DataFrame
        # ------------------------------------------

        self.build_dataframe()

        # ------------------------------------------
        # Runtime state
        # ------------------------------------------

        self.modified = False

        self.loaded = True

        if len(self.df.columns):

            self.product_column = self.df.columns[-1]

        else:

            self.product_column = None

    # ==========================================================
    # Reload
    # ==========================================================

    def reload(self):

        if not self.loaded:

            return False

        if not self.filename:

            return False

        self.load_local(

            self.filename

        )

        return True

    # ==========================================================
    # Switch Document
    # ==========================================================

    def switch_document(

        self,

        document_name

    ):

        if self.loaded:

            self.close()

        self.open(

            document_name

        )

        return True

    # ==========================================================
    # Close
    # ==========================================================

    def close(self):

        if not self.loaded:

            return

        try:

            if self.online:

                self.synchronize()

            else:

                self.save_local()

        finally:

            self.network.disconnect()

            self.reset_runtime()

            self.loaded = False

    def synchronize(self):

        pass

    def reset_runtime(self):

        pass

    # ==========================================================
    # Workbook Metadata
    # ==========================================================

    def load_metadata(self):

        sheet = self.sheet

        # ----------------------------
        # Merged Cells
        # ----------------------------

        self.merged_cells = [

            str(rng)

            for rng in sheet.merged_cells.ranges

        ]

        # ----------------------------
        # Column Widths
        # ----------------------------

        self.column_widths = {}

        for key, dimension in sheet.column_dimensions.items():

            self.column_widths[key] = dimension.width

        # ----------------------------
        # Row Heights
        # ----------------------------

        self.row_heights = {}

        for key, dimension in sheet.row_dimensions.items():

            self.row_heights[key] = dimension.height

        # ----------------------------
        # Hidden Rows
        # ----------------------------

        self.hidden_rows = {

            row

            for row, dimension

            in sheet.row_dimensions.items()

            if dimension.hidden

        }

        # ----------------------------
        # Hidden Columns
        # ----------------------------

        self.hidden_columns = {

            column

            for column, dimension

            in sheet.column_dimensions.items()

            if dimension.hidden

        }

        # ----------------------------
        # Freeze Panes
        # ----------------------------

        self.freeze_panes = sheet.freeze_panes

        # ----------------------------
        # Auto Filter
        # ----------------------------

        self.auto_filter = sheet.auto_filter.ref

    # ==========================================================
    # Build DataFrame
    # ==========================================================

    def build_dataframe(self):

        self.df = self.storage.worksheet_to_dataframe(

            self.sheet

        )

        self.filtered_df = self.df.copy()

        return self.df

    # ==========================================================
    # Refresh Views
    # ==========================================================

    def refresh_views(self):

        self.build_dataframe()

        # Future

        # self.build_view_model()

        return True

    # ==========================================================
    # Sheet Management
    # ==========================================================

    def list_sheets(self):

        self.sheet_names = self.workbook.sheetnames

        return self.sheet_names.copy()

    def get_sheet(self, sheet_name):

        if sheet_name not in self.workbook.sheetnames:

            return None

        return self.workbook[sheet_name]

    # ==========================================================
    # Set Sheet
    # ==========================================================

    def set_sheet(self, sheet_name):

        if self.sheet_name == sheet_name:

            return True

        worksheet = self.get_sheet(

            sheet_name

        )

        if worksheet is None:

            return False

        self.sheet = worksheet

        self.sheet_name = sheet_name

        self.active_sheet_index = (

            self.workbook.sheetnames.index(

                sheet_name

            )

        )

        self.load_metadata()

        self.refresh_views()

        return True

    # ==========================================================
    # Create Sheet
    # ==========================================================

    def create_sheet(

        self,

        sheet_name,

        activate=True

    ):

        if sheet_name in self.workbook.sheetnames:

            return False

        self.workbook.create_sheet(

            title=sheet_name

        )

        self.list_sheets()

        if activate:

            self.set_sheet(

                sheet_name

            )

        return True

    # ==========================================================
    # Duplicate Sheet
    # ==========================================================

    def duplicate_sheet(

        self,

        source_name,

        new_name=None,

        activate=True

    ):

        source = self.get_sheet(

            source_name

        )

        if source is None:

            return False

        copy = self.workbook.copy_worksheet(

            source

        )

        if new_name is None:

            base = source_name + " Copy"

            new_name = base

            counter = 2

            while new_name in self.workbook.sheetnames:

                new_name = f"{base} {counter}"

                counter += 1

        elif new_name in self.workbook.sheetnames:

            return False

        copy.title = new_name

        self.list_sheets()

        if activate:

            self.set_sheet(

                new_name

            )

        return True

    # ==========================================================
    # Delete Sheet
    # ==========================================================

    def delete_sheet(

        self,

        sheet_name

    ):

        worksheet = self.get_sheet(

            sheet_name

        )

        if worksheet is None:

            return False

        if len(self.workbook.sheetnames) <= 1:

            return False

        was_active = (

            self.sheet_name == sheet_name

        )

        next_sheet = None

        if was_active:

            names = self.workbook.sheetnames

            index = names.index(sheet_name)

            if index > 0:

                next_sheet = names[index - 1]

            else:

                next_sheet = names[1]

        self.workbook.remove(

            worksheet

        )

        self.list_sheets()

        if was_active:

            self.set_sheet(

                next_sheet

            )

        return True

