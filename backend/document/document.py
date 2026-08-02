from typing import Any

import pandas as pd
from openpyxl import Workbook


from backend.models.spreadsheet_status import DocumentInfo, SpreadsheetData, SpreadsheetSheet
from backend.utils.config import DOCUMENTS_FOLDER
from backend.utils.logger import get_logger

from backend.models.spreadsheet import Spreadsheet
from backend.models.sheet import Sheet
from backend.models.spreadsheet_row import SpreadsheetRow

from backend.services.search_service import SearchService
from backend.services.sort_service import SortService
from backend.services.editing_service import EditingService

logger = get_logger("document")


class Document:
    """Own spreadsheet state and workbook operations for the backend."""

    def __init__(self, storage: Any, network: Any, online: bool = False) -> None:
        """
        Initialize the document state and its collaborators.

        Args:
            storage: Storage backend used for file operations.
            network: Network backend placeholder for future sync support.
            online: Whether the document is connected to a remote service.
        """
        self.storage = storage
        self.network = network
        self.online = online

        self.workbook = None
        self.sheet = None

        self.filename = ""
        self.sheet_name = ""

        self.search_service = SearchService()
        self.sort_service = SortService()
        self.editing_service = EditingService()

        self.documents_folder = DOCUMENTS_FOLDER

        self.df = pd.DataFrame()
        self.filtered_df = pd.DataFrame()

        # Domain Model
        self.spreadsheet = Spreadsheet()

        self.search_text = ""
        self.sort_rules: list[dict[str, Any]] = []

        self.loaded = False
        self.modified = False

        self.undo_stack: list[dict[str, Any]] = []
        self.redo_stack: list[dict[str, Any]] = []

    # ==========================================================
    # Document
    # ==========================================================

    def open(self, filename: str) -> bool:
        """
        Open an Excel document from disk.

        Args:
            filename: Path to the workbook.

        Returns:
            True if the document was loaded successfully, otherwise False.
        """
        if not self._is_valid_filename(filename):
            logger.warning("Open rejected for invalid filename: %s", filename)
            return False

        if not self.load_local(filename):
            logger.warning("Failed to open workbook: %s", filename)
            return False

        self.filename = filename
        self.loaded = True
        self.modified = False
        logger.info("Opened document %s", filename)
        return True

    def create(self, filename: str) -> bool:
        """
        Create a new Excel document.

        Args:
            filename: Path for the new workbook.

        Returns:
            True if the new document was created successfully, otherwise False.
        """
        if not self._is_valid_filename(filename):
            logger.warning("Create rejected for invalid filename: %s", filename)
            return False

        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Sheet1"

        self.filename = filename
        self.sheet_name = self.sheet.title

        self._clear_view_data()

        self.loaded = True
        self.modified = True
        logger.info("Created document %s", filename)
        return True

    def close(self) -> bool:
        """
        Close the active document and clear its state.

        Returns:
            True if the document was closed successfully, otherwise False.
        """
        self._reset_state()
        return True

    def reload(self) -> bool:
        """
        Reload the active document from disk.

        Returns:
            True if the document was reloaded successfully, otherwise False.
        """
        if not self.loaded or not self.filename:
            return False

        return self.load_local(self.filename)


    # ==========================================================
    # Domain Model
    # ==========================================================

    def dataframe_to_spreadsheet(self) -> None:

        if self.sheet is None:
        
            self.spreadsheet = Spreadsheet()

            return


        rows = []

        for _, row in self.df.iterrows():

            values = row.fillna("").to_dict()

            rows.append(

                SpreadsheetRow(values=values)

            )


        sheet = Sheet(

            name=self.sheet.title,

            rows=rows
        )


        self.spreadsheet = Spreadsheet(

            sheets=[sheet]
        )


    # ==========================================================
    # Loading / Saving
    # ==========================================================

    def load_local(self, filename: str) -> bool:
        """
        Load a workbook from storage into memory.

        Args:
            filename: Path to the workbook.

        Returns:
            True if the workbook was loaded successfully, otherwise False.
        """
        if not self._is_valid_filename(filename):
            return False

        self.workbook = self.storage.open(filename)

        if self.workbook is None:
            return False

        self.sheet = self.workbook.active
        self.sheet_name = self.sheet.title

        self._set_view_data(self.storage.read_sheet(self.sheet))
        self.dataframe_to_spreadsheet()
        self.modified = False
        return True

    def save_local(self) -> bool:
        """
        Save the active workbook to disk.

        Returns:
            True if the workbook was saved successfully, otherwise False.
        """
        if self.workbook is None or not self.filename:
            logger.warning("Save skipped because no active workbook is available")
            return False

        self.storage.write_sheet(self.sheet, self.df)
        saved = self.storage.save(self.workbook, self.filename)
        self.modified = False

        if saved:
            logger.info("Saved document %s", self.filename)
        else:
            logger.warning("Failed to save document %s", self.filename)

        return saved

    # ==========================================================
    # Search
    # ==========================================================

    def search(self, text: str) -> bool:
        """
        Filter the visible rows using a search query.

        Args:
            text: Text to search for.

        Returns:
            True if the search completed successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        self.search_text = text

        self.filtered_df = self.search_service.search(

            self.df,

            text

        )

        if self.sort_rules:

            self.sort(

                self.sort_rules,

                reapply=True

            )

        return True

    def clear_search(self) -> bool:
        """
        Clear the active search filter.

        Returns:
            True if the search filter was cleared successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        self.search_text = ""
        self.filtered_df = self.df.copy()
        return self._reapply_sort()

    # ==========================================================
    # Sorting
    # ==========================================================

    def sort(self, sort_rules: list[dict[str, Any]], reapply: bool = False) -> bool:
        """
        Apply sort rules to the filtered view.

        Args:
            sort_rules: A list of sort rule dictionaries.
            reapply: Whether to reuse the current sort rules.

        Returns:
            True if sorting completed successfully, otherwise False.
        """
        if not reapply:
            self.sort_rules = sort_rules

        self.filtered_df = self.sort_service.sort(

            self.filtered_df,

            self.sort_rules

        )

        return True

    def clear_sort(self) -> bool:
        """
        Clear the active sort rules.

        Returns:
            True if sorting was reset successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        self.sort_rules = []
        self.filtered_df = self.df.copy()
        return True

    # ==========================================================
    # Editing
    # ==========================================================

    def _sync_active_sheet(self) -> None:
        """Write the current DataFrame to the active worksheet."""
        if self.workbook is None or self.sheet is None:
            return

        self.storage.write_sheet(self.sheet, self.df)

    def _refresh_view(self) -> None:
        """Reload the active worksheet and reapply the current search/sort state."""
        if self.workbook is None or self.sheet is None:
            self._clear_view_data()
            return

        self._set_view_data(self.storage.read_sheet(self.sheet), apply_search=True)

    def edit_cell(self, row: int, column: int, value: Any) -> bool:
        """
        Update a cell value in the active document.

        Args:
            row: Zero-based row index.
            column: Zero-based column index.
            value: New value for the cell.

        Returns:
            True if the update completed successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if row < 0 or column < 0:
            return False

        if row >= len(self.df):
            return False

        if column >= len(self.df.columns):
            return False

        success = self.editing_service.edit_cell(

            self.spreadsheet,

            row,

            column,

            value

        )
        if not success:
            return False

        self.df.iloc[row, column] = value
        self._sync_active_sheet()

        self.modified = True

        self.search(self.search_text)

        return True

    def insert_row(self, index: int | None = None) -> bool:
        """
        Insert a new row into the active document.

        Args:
            index: Optional zero-based index where the row should be inserted.

        Returns:
            True if the row was inserted successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if index is None:
            index = len(self.df)

        if index < 0 or index > len(self.df):
            return False

        empty = {column: "" for column in self.df.columns}
        top = self.df.iloc[:index]
        bottom = self.df.iloc[index:]

        self.df = pd.concat(
            [top, pd.DataFrame([empty]), bottom],
            ignore_index=True
        )
        self._sync_active_sheet()
        self.modified = True
        self.search(self.search_text)
        return True

    def delete_row(self, index: int) -> bool:
        """
        Delete a row from the active document.

        Args:
            index: Zero-based row index to remove.

        Returns:
            True if the row was deleted successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if index < 0 or index >= len(self.df):
            return False

        self.df = self.df.drop(index)
        self.df.reset_index(drop=True, inplace=True)
        self._sync_active_sheet()
        self.modified = True
        self.search(self.search_text)
        return True

    def insert_column(self, name: str, index: int | None = None) -> bool:
        """
        Insert a new column into the active document.

        Args:
            name: Name of the new column.
            index: Optional zero-based index for the new column.

        Returns:
            True if the column was inserted successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if not self._is_valid_filename(name):
            return False

        if name in self.df.columns:
            return False

        if index is None:
            index = len(self.df.columns)

        if index < 0 or index > len(self.df.columns):
            return False

        self.df.insert(index, name, "")
        self._sync_active_sheet()
        self.modified = True
        self.search(self.search_text)
        return True

    def delete_column(self, name: str) -> bool:
        """
        Delete a column from the active document.

        Args:
            name: Column name to remove.

        Returns:
            True if the column was deleted successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if name not in self.df.columns:
            return False

        self.df.drop(columns=[name], inplace=True)
        self._sync_active_sheet()
        self.modified = True
        self.search(self.search_text)
        return True

    def rename_sheet(self, old_name: str, new_name: str) -> bool:
        """
        Rename a worksheet in the active document.

        Args:
            old_name: Current worksheet name.
            new_name: New worksheet name.

        Returns:
            True if the worksheet was renamed successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if not self._is_valid_filename(old_name) or not self._is_valid_filename(new_name):
            return False

        if old_name not in self.workbook.sheetnames:
            return False

        if new_name in self.workbook.sheetnames and new_name != old_name:
            return False

        self.workbook[old_name].title = new_name
        self.sheet_name = new_name
        self.sheet = self.workbook[new_name]
        self._sync_active_sheet()
        self.modified = True
        return True

    def add_sheet(self, name: str) -> bool:
        """
        Add a new worksheet to the active document.

        Args:
            name: Name of the new worksheet.

        Returns:
            True if the worksheet was added successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if not self._is_valid_filename(name):
            return False

        if name in self.workbook.sheetnames:
            return False

        self.workbook.create_sheet(title=name)
        self.sheet = self.workbook[name]
        self.sheet_name = name
        self._clear_view_data()
        self._sync_active_sheet()
        self.modified = True
        self.search(self.search_text)
        return True

    def delete_sheet(self, name: str) -> bool:
        """
        Delete a worksheet from the active document.

        Args:
            name: Worksheet name to remove.

        Returns:
            True if the worksheet was deleted successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if not self._is_valid_filename(name):
            return False

        if name not in self.workbook.sheetnames:
            return False

        if len(self.workbook.sheetnames) == 1:
            return False

        current_sheet_removed = name == self.sheet_name

        self.workbook.remove(self.workbook[name])
        self.modified = True

        if current_sheet_removed:
            remaining_sheet = self.workbook.sheetnames[0]
            self.sheet = self.workbook[remaining_sheet]
            self.sheet_name = remaining_sheet
            self._refresh_view()
            return True

        if self.sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[self.sheet_name]
        else:
            self.sheet = self.workbook.active
            self.sheet_name = self.sheet.title

        self._refresh_view()
        return True

    # ==========================================================
    # Sheets
    # ==========================================================

    def list_sheets(self) -> list[str]:
        """
        Return the worksheet names for the active workbook.

        Returns:
            A list of worksheet names.
        """
        if self.workbook is None:
            return []

        return self.workbook.sheetnames

    def set_sheet(self, sheet_name: str) -> bool:
        """
        Switch to a different worksheet.

        Args:
            sheet_name: The worksheet name to activate.

        Returns:
            True if the worksheet changed successfully, otherwise False.
        """
        if self.workbook is None:
            return False

        if sheet_name not in self.workbook.sheetnames:
            return False

        self.sheet = self.workbook[sheet_name]
        self.sheet_name = sheet_name

        self._set_view_data(self.storage.read_sheet(self.sheet), apply_search=True)
        return True

    # ==========================================================
    # File Management
    # ==========================================================

    def list_documents(self, folder: str) -> list[str]:
        """
        List Excel documents in a folder.

        Args:
            folder: Folder to inspect.

        Returns:
            A sorted list of workbook filenames.
        """
        return self.storage.list_documents(folder)

    def delete_document(self, filename: str) -> bool:
        """
        Delete an Excel document from disk.

        Args:
            filename: Path to the workbook.

        Returns:
            True if the workbook was deleted successfully, otherwise False.
        """
        if not self._is_valid_filename(filename):
            logger.warning("Delete rejected for invalid filename: %s", filename)
            return False

        deleted = self.storage.delete(filename)

        if deleted:
            logger.info("Deleted document %s", filename)
        else:
            logger.warning("Failed to delete document %s", filename)

        return deleted

    def copy_document(self, source: str, destination: str) -> bool:
        """
        Copy an Excel document to a new location.

        Args:
            source: Source workbook path.
            destination: Destination workbook path.

        Returns:
            True if the workbook was copied successfully, otherwise False.
        """
        if not self._is_valid_filename(source) or not self._is_valid_filename(destination):
            return False
        return self.storage.copy(source, destination)

    def rename_document(self, old_name: str, new_name: str) -> bool:
        """
        Rename an Excel document on disk.

        Args:
            old_name: Current workbook path.
            new_name: New workbook path.

        Returns:
            True if the workbook was renamed successfully, otherwise False.
        """
        if not self._is_valid_filename(old_name) or not self._is_valid_filename(new_name):
            return False

        return self.storage.rename(old_name, new_name)

    # ==========================================================
    # Helpers
    # ==========================================================

    def table(self) -> pd.DataFrame:
        """
        Return the currently filtered table view.

        Returns:
            The filtered pandas DataFrame.
        """
        return self.filtered_df

    def data(self) -> SpreadsheetData:
        """
        Return the currently filtered spreadsheet data as a structured value object.

        Returns:
            A structured representation of the visible rows.
        """
        return SpreadsheetData(
            headers=self.headers(),
            rows=self.filtered_df.fillna("").to_dict("records"),
            row_count=self.row_count(),
            column_count=self.column_count()
        )

    def original_table(self) -> pd.DataFrame:
        """
        Return the underlying document table.

        Returns:
            The full pandas DataFrame.
        """
        return self.df

    def filtered_row_count(self) -> int:
        """Return the number of rows in the filtered view."""
        return len(self.filtered_df)

    def row_count(self) -> int:
        """
        Return the number of rows in the filtered view.

        Returns:
            The filtered row count.
        """

        sheet = self.spreadsheet.current_sheet()

        if sheet is None:
            return 0

        return sheet.row_count

    def column_count(self) -> int:
        """
        Return the number of columns in the filtered view.

        Returns:
            The filtered column count.
        """
        sheet = self.spreadsheet.current_sheet()

        if sheet is None:
            return 0

        return sheet.column_count

    def headers(self) -> list[str]:
        """
        Return the visible headers.

        Returns:
            The visible column names.
        """
        sheet = self.spreadsheet.current_sheet()

        if sheet is None:
            return []

        return sheet.headers

    def info(self) -> DocumentInfo:
        """
        Return a structured summary of the active document.

        Returns:
            A document summary object.
        """
        return DocumentInfo(
            filename=self.filename,
            loaded=self.loaded,
            modified=self.modified,
            rows=self.row_count(),
            columns=self.column_count(),
            sheets=self.list_sheets(),
            current_sheet=self.sheet_name
        )

    def sheet_metadata(self) -> list[SpreadsheetSheet]:
        """
        Return worksheet metadata for the active workbook.

        Returns:
            A list of worksheet metadata objects.
        """
        if self.workbook is None:
            return []

        return [
            SpreadsheetSheet(name=name, active=name == self.sheet_name)
            for name in self.workbook.sheetnames
        ]

    def _reset_state(self) -> None:
        """Reset the workbook and document state."""
        self.workbook = None
        self.sheet = None
        self.filename = ""
        self.sheet_name = ""
        self._clear_view_data()
        self.search_text = ""
        self.sort_rules = []
        self.loaded = False
        self.modified = False

    def _reapply_sort(self) -> bool:
        """Reapply the current sort rules to the filtered view."""
        if self.sort_rules:
            return self.sort(self.sort_rules, reapply=True)
        return True

    def _is_valid_filename(self, filename: str | None) -> bool:
        """Validate that a filename is a non-empty string."""
        return isinstance(filename, str) and bool(filename.strip())

    def _clear_view_data(self) -> None:
        """Reset the in-memory DataFrame view state."""
        self.df = pd.DataFrame()
        self.filtered_df = pd.DataFrame()

    def _set_view_data(self, dataframe: pd.DataFrame, apply_search: bool = False) -> None:
        """Populate the active DataFrame view and optionally reapply search state."""
        self.df = dataframe
        self.filtered_df = dataframe.copy()

        if apply_search:
            self.search(self.search_text)

    def spreadsheet_to_dataframe(self) -> None:

        sheet = self.spreadsheet.current_sheet()

        if sheet is None:

            self._clear_view_data()

            return

        rows = [

            row.values

            for row in sheet.rows

        ]

        self._set_view_data(pd.DataFrame(rows))

    