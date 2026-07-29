import os
import tempfile

from openpyxl import Workbook

from backend.controller.controller import Controller


def test_controller_data_and_sheet_switching():
    with tempfile.TemporaryDirectory() as tmp_dir:
        workbook_path = os.path.join(tmp_dir, "sample.xlsx")

        workbook = Workbook()
        sheet_one = workbook.active
        sheet_one.title = "Sheet1"
        sheet_one["A1"] = "name"
        sheet_one["A2"] = "Alice"

        sheet_two = workbook.create_sheet("Sheet2")
        sheet_two["A1"] = "name"
        sheet_two["A2"] = "Bob"

        workbook.save(workbook_path)

        controller = Controller()

        assert controller.open_document(workbook_path) is True
        assert controller.headers() == ["name"]
        assert controller.data() == [{"name": "Alice"}]
        assert controller.sheets() == ["Sheet1", "Sheet2"]
        assert controller.current_sheet() == "Sheet1"
        assert controller.set_sheet("Sheet2") is True
        assert controller.current_sheet() == "Sheet2"
        assert controller.data() == [{"name": "Bob"}]


def test_controller_search_and_clear_search_return_success_values():
    with tempfile.TemporaryDirectory() as tmp_dir:
        workbook_path = os.path.join(tmp_dir, "search.xlsx")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet["A1"] = "name"
        sheet["A2"] = "Alice"
        sheet["A3"] = "Bob"

        workbook.save(workbook_path)

        controller = Controller()

        assert controller.open_document(workbook_path) is True
        assert controller.search("Alice") is True
        assert controller.data() == [{"name": "Alice"}]
        assert controller.clear_search() is True
        assert controller.data() == [{"name": "Alice"}, {"name": "Bob"}]


def test_controller_editing_operations_update_dataframe_and_workbook():
    with tempfile.TemporaryDirectory() as tmp_dir:
        workbook_path = os.path.join(tmp_dir, "editing.xlsx")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet["A1"] = "name"
        sheet["A2"] = "Alice"
        sheet["A3"] = "Bob"
        workbook.save(workbook_path)

        controller = Controller()

        assert controller.open_document(workbook_path) is True

        assert controller.edit_cell(0, 0, "Updated") is True
        assert controller.data() == [{"name": "Updated"}, {"name": "Bob"}]
        assert controller.loaded_document().sheet["A2"].value == "Updated"

        assert controller.insert_row(1) is True
        assert controller.data() == [{"name": "Updated"}, {"name": ""}, {"name": "Bob"}]

        assert controller.delete_row(1) is True
        assert controller.data() == [{"name": "Updated"}, {"name": "Bob"}]

        assert controller.insert_column("age", 1) is True
        assert controller.data() == [
            {"name": "Updated", "age": ""},
            {"name": "Bob", "age": ""}
        ]

        assert controller.delete_column("age") is True
        assert controller.data() == [{"name": "Updated"}, {"name": "Bob"}]

        assert controller.add_sheet("Sheet2") is True
        assert controller.sheets() == ["Sheet1", "Sheet2"]

        assert controller.rename_sheet("Sheet2", "Sheet2Renamed") is True
        assert controller.sheets() == ["Sheet1", "Sheet2Renamed"]

        assert controller.set_sheet("Sheet2Renamed") is True
        assert controller.delete_sheet("Sheet2Renamed") is True
        assert controller.current_sheet() == "Sheet1"


def test_controller_failure_cases_return_false():
    with tempfile.TemporaryDirectory() as tmp_dir:
        workbook_path = os.path.join(tmp_dir, "failures.xlsx")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet["A1"] = "name"
        sheet["A2"] = "Alice"
        workbook.save(workbook_path)

        controller = Controller()

        assert controller.open_document(workbook_path) is True
        assert controller.edit_cell(1, 0, "X") is False
        assert controller.edit_cell(0, 3, "X") is False
        assert controller.insert_row(-1) is False
        assert controller.delete_row(10) is False
        assert controller.insert_column("name") is False
        assert controller.add_sheet("Sheet1") is False
        assert controller.rename_sheet("Missing", "Other") is False
