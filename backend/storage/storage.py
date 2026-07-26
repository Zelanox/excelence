import copy
import os
import tempfile

import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================================
# Initialization
# ==========================================================

class ExcelStorage:

    def __init__(
        self,
        read_only=False,
        keep_vba=False,
        data_only=False,
    ):

        self.read_only = read_only
        self.keep_vba = keep_vba
        self.data_only = data_only

    # ==========================================================
    # Workbook
    # ==========================================================

    def open(self, filename):

        if not self.exists(filename):
            return None

        if not self.is_excel_file(filename):
            return None

        return load_workbook(
            filename,
            read_only=self.read_only,
            keep_vba=self.keep_vba,
            data_only=self.data_only
        )


    def create(self):

        return Workbook()


    def save(self, workbook, filename):

        if workbook is None:
            return False

        folder = os.path.dirname(filename)

        if folder:
            os.makedirs(folder, exist_ok=True)

        fd, temp = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)

        try:

            workbook.save(temp)

            os.replace(
                temp,
                filename
            )

        finally:

            if os.path.exists(temp):

                try:
                    os.remove(temp)
                except Exception:
                    pass

        return True

    # ==========================================================
    # Worksheet <-> DataFrame
    # ==========================================================

    def read_sheet(self, sheet):

        rows = list(sheet.values)

        if not rows:
            return pd.DataFrame()

        dataframe = pd.DataFrame(
            rows[1:],
            columns=rows[0]
        )

        return self._normalize_dataframe(dataframe)


    def write_sheet(self, sheet, dataframe):

        self._clear_sheet(sheet)

        for row in dataframe_to_rows(
            dataframe,
            index=False,
            header=True
        ):
            sheet.append(row)

        return sheet


    # ==========================================================
    # Private Helpers
    # ==========================================================

    def _normalize_dataframe(self, dataframe):

        dataframe = dataframe.copy()

        dataframe.columns = [
            str(column)
            for column in dataframe.columns
        ]

        return dataframe


    def _clear_sheet(self, sheet):

        if sheet.max_row:
            sheet.delete_rows(
                1,
                sheet.max_row
            )

    # ==========================================================
    # Extract Styles
    # ==========================================================

    def _extract_styles(self, worksheet):

        styles = {

            "column_dimensions": (
                self._extract_column_dimensions(
                    worksheet
                )
            ),

            "row_dimensions": (
                self._extract_row_dimensions(
                    worksheet
                )
            ),

            "cell_styles": (
                self._extract_cell_styles(
                    worksheet
                )
            )

        }

        return styles

    # ==========================================================
    # Extract Column Dimensions
    # ==========================================================

    def _extract_column_dimensions(self, worksheet):

        dimensions = {}

        for column, dimension in (
            worksheet.column_dimensions.items()
        ):

            dimensions[column] = {

                "width": dimension.width,

                "hidden": dimension.hidden

            }

        return dimensions

    # ==========================================================
    # Extract Row Dimensions
    # ==========================================================

    def _extract_row_dimensions(self, worksheet):

        dimensions = {}

        for row, dimension in (
            worksheet.row_dimensions.items()
        ):

            dimensions[row] = {

                "height": dimension.height,

                "hidden": dimension.hidden

            }

        return dimensions

    # ==========================================================
    # Extract Cell Styles
    # ==========================================================

    def _extract_cell_styles(self, worksheet):

        styles = {}

        for row in worksheet.iter_rows():

            for cell in row:

                if cell.has_style:

                    styles[cell.coordinate] = {

                        "font": copy.copy(cell.font),

                        "fill": copy.copy(cell.fill),

                        "border": copy.copy(cell.border),

                        "alignment": copy.copy(cell.alignment),

                        "number_format": cell.number_format,

                        "protection": copy.copy(cell.protection)

                    }

        return styles

    # ==========================================================
    # Restore Styles
    # ==========================================================

    def _restore_styles(self, worksheet, styles):

        if not styles:
            return

        self._restore_column_dimensions(

            worksheet,

            styles.get(
                "column_dimensions",
                {}
            )

        )

        self._restore_row_dimensions(

            worksheet,

            styles.get(
                "row_dimensions",
                {}
            )

        )

        self._restore_cell_styles(

            worksheet,

            styles.get(
                "cell_styles",
                {}
            )

        )

    # ==========================================================
    # Restore Column Dimensions
    # ==========================================================

    def _restore_column_dimensions(
        self,
        worksheet,
        dimensions
    ):

        for column, data in dimensions.items():

            dimension = worksheet.column_dimensions[column]

            dimension.width = data.get("width")

            dimension.hidden = data.get(
                "hidden",
                False
            )

    # ==========================================================
    # Restore Row Dimensions
    # ==========================================================

    def _restore_row_dimensions(
        self,
        worksheet,
        dimensions
    ):

        for row, data in dimensions.items():

            dimension = worksheet.row_dimensions[row]

            dimension.height = data.get("height")

            dimension.hidden = data.get(
                "hidden",
                False
            )

    # ==========================================================
    # Restore Cell Styles
    # ==========================================================

    def _restore_cell_styles(
        self,
        worksheet,
        styles
    ):

        for coordinate, style in styles.items():

            cell = worksheet[coordinate]

            cell.font = copy.copy(
                style["font"]
            )

            cell.fill = copy.copy(
                style["fill"]
            )

            cell.border = copy.copy(
                style["border"]
            )

            cell.alignment = copy.copy(
                style["alignment"]
            )

            cell.number_format = (
                style["number_format"]
            )

            cell.protection = copy.copy(
                style["protection"]
            )

