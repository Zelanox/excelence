import copy
import os
import tempfile
import shutil

import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================================
# Initialization
# ==========================================================

class ExcelStorage:

    def __init__(

        self,

        read_only=False,

        keep_vba=False,

        data_only=False

    ):

        self.read_only = read_only

        self.keep_vba = keep_vba

        self.data_only = data_only

    # ==========================================================
    # Workbook Validation
    # ==========================================================

    def validate_workbook(

        self,

        filename

    ):

        return (

            self.workbook_exists(filename)

            and

            self.validate_extension(filename)

        )

    # ==========================================================
    # Open Workbook
    # ==========================================================

    def open_workbook(

        self,

        filename

    ):

        if not self.validate_workbook(

            filename

        ):

            return None

        workbook = load_workbook(

            filename,

            read_only=self.read_only,

            keep_vba=self.keep_vba,

            data_only=self.data_only

        )

        return workbook

    # ==========================================================
    # Create Workbook
    # ==========================================================

    def create_workbook(

        self

    ):

        workbook = Workbook()

        return workbook

    # ==========================================================
    # Save Workbook
    # ==========================================================

    def save_workbook(

        self,

        workbook,

        filename=None

    ):

        if filename is None:

            return False

        folder = os.path.dirname(filename)

        if folder:

            os.makedirs(

                folder,

                exist_ok=True

            )

        fd, temp_name = tempfile.mkstemp(

            suffix=".xlsx"

        )

        os.close(fd)

        try:

            workbook.save(

                temp_name

            )

            os.replace(

                temp_name,

                filename

            )

        except Exception:

            if os.path.exists(temp_name):

                os.remove(temp_name)

            raise

        return True

    def save_as_workbook(

        self,

        workbook,

        filename

    ):

        return self.save_workbook(

            workbook,

            filename

        )

    # ==========================================================
    # Worksheet -> DataFrame
    # ==========================================================

    def worksheet_to_dataframe(

        self,

        worksheet

    ):

        rows = list(

            worksheet.values

        )

        if not rows:

            return pd.DataFrame()

        dataframe = pd.DataFrame(

            rows[1:],

            columns=rows[0]

        )

        return self.normalize_dataframe(

            dataframe

        )

    # ==========================================================
    # Normalize DataFrame
    # ==========================================================

    def normalize_dataframe(

        self,

        dataframe

    ):

        dataframe = dataframe.copy()

        dataframe.columns = [

            str(column)

            for column in dataframe.columns

        ]

        return dataframe

    # ==========================================================
    # Clear Worksheet
    # ==========================================================

    def clear_worksheet(

        self,

        worksheet

    ):

        worksheet.delete_rows(

            1,

            worksheet.max_row

        )

    # ==========================================================
    # DataFrame -> Worksheet
    # ==========================================================

    def dataframe_to_worksheet(

        self,

        dataframe,

        worksheet

    ):

        self.clear_worksheet(

            worksheet

        )

        for row in dataframe_to_rows(

            dataframe,

            index=False,

            header=True

        ):

            worksheet.append(

                row

            )

        return worksheet

    # ==========================================================
    # Extract Metadata
    # ==========================================================

    def extract_metadata(

        self,

        worksheet

    ):

        metadata = {

            "merged_cells": self.extract_merged_cells(

                worksheet

            ),

            "freeze_panes": None,

            "auto_filter": None,

            "row_dimensions": None,

            "column_dimensions": None

        }

        return metadata

    # ==========================================================
    # Extract Merged Cells
    # ==========================================================

    def extract_merged_cells(

        self,

        worksheet

    ):

        return [

            str(range_)

            for range_

            in worksheet.merged_cells.ranges

        ]

    # ==========================================================
    # Restore Metadata
    # ==========================================================

    def restore_metadata(

        self,

        worksheet,

        metadata

    ):

        self.restore_merged_cells(

            worksheet,

            metadata.get(

                "merged_cells",

                []

            )

        )

    # ==========================================================
    # Restore Merged Cells
    # ==========================================================

    def restore_merged_cells(

        self,

        worksheet,

        merged_ranges

    ):

        for cell_range in merged_ranges:

            worksheet.merge_cells(

                cell_range

            )

    # ==========================================================
    # Extract Styles
    # ==========================================================

    def extract_styles(

        self,

        worksheet

    ):

        styles = {

            "column_dimensions": (

                self.extract_column_dimensions(

                    worksheet

                )

            ),

            "row_dimensions": (

                self.extract_row_dimensions(

                    worksheet

                )

            ),

            "cell_styles": (

                self.extract_cell_styles(

                    worksheet

                )

            )
        }

        return styles

    # ==========================================================
    # Extract Column Dimensions
    # ==========================================================

    def extract_column_dimensions(

        self,

        worksheet

    ):

        dimensions = {}

        for column, dimension in (

            worksheet.column_dimensions.items()

        ):

            dimensions[column] = {

                "width": dimension.width,

                "hidden": dimension.hidden

            }

        return dimensions

    # ==========================================================
    # Extract Row Dimensions
    # ==========================================================

    def extract_row_dimensions(

        self,

        worksheet

    ):

        dimensions = {}

        for row, dimension in (

            worksheet.row_dimensions.items()

        ):

            dimensions[row] = {

                "height": dimension.height,

                "hidden": dimension.hidden

            }

        return dimensions

    # ==========================================================
    # Restore Styles
    # ==========================================================

    def restore_styles(

        self,

        worksheet,

        styles

    ):

        if not styles:

            return

        self.restore_column_dimensions(

            worksheet,

            styles.get(

                "column_dimensions",

                {}

            )

        )

        self.restore_row_dimensions(

            worksheet,

            styles.get(

                "row_dimensions",

                {}

            )

        )

        self.restore_cell_styles(

            worksheet,

            styles.get(

                "cell_styles",

                {}

            )

        )



    # ==========================================================
    # Restore Column Dimensions
    # ==========================================================

    def restore_column_dimensions(

        self,

        worksheet,

        dimensions

    ):

        for column, data in dimensions.items():

            dimension = worksheet.column_dimensions[column]

            dimension.width = data.get(

                "width"

            )

            dimension.hidden = data.get(

                "hidden",

                False

            )

    # ==========================================================
    # Restore Row Dimensions
    # ==========================================================

    def restore_row_dimensions(

        self,

        worksheet,

        dimensions

    ):

        for row, data in dimensions.items():

            dimension = worksheet.row_dimensions[row]

            dimension.height = data.get(

                "height"

            )

            dimension.hidden = data.get(

                "hidden",

                False

            )

    # ==========================================================
    # Extract Cell Styles
    # ==========================================================

    def extract_cell_styles(

        self,

        worksheet

    ):

        styles = {}

        for row in worksheet.iter_rows():

            for cell in row:

                if cell.has_style:

                    styles[cell.coordinate] = {

                        "font": copy.copy(

                            cell.font

                        ),

                        "fill": copy.copy(

                            cell.fill

                        ),

                        "border": copy.copy(

                            cell.border

                        ),

                        "alignment": copy.copy(

                            cell.alignment

                        ),

                        "number_format": (

                            cell.number_format

                        ),

                        "protection": copy.copy(

                            cell.protection

                        )

                    }

        return styles

    # ==========================================================
    # Restore Cell Styles
    # ==========================================================

    def restore_cell_styles(

        self,

        worksheet,

        styles

    ):

        for coordinate, style in styles.items():

            cell = worksheet[coordinate]

            cell.font = copy.copy(

                style["font"]

            )

            cell.fill = copy.copy(

                style["fill"]

            )

            cell.border = copy.copy(

                style["border"]

            )

            cell.alignment = copy.copy(

                style["alignment"]

            )

            cell.number_format = (

                style["number_format"]

            )

            cell.protection = copy.copy(

                style["protection"]

            )

    # ==========================================================
    # File Services
    # ==========================================================

    def workbook_exists(

        self,

        filename

    ):

        return os.path.exists(

            filename

        )

    # ==========================================================
    # Validators
    # ==========================================================

    def validate_extension(

        self,

        filename

    ):

        extension = os.path.splitext(

            filename

        )[1].lower()

        return extension in (

            ".xlsx",

            ".xlsm",

            ".xltx",

            ".xltm"

        )

    # ==========================================================
    # Backup
    # ==========================================================

    def backup_workbook(

        self,

        filename

    ):

        if not os.path.exists(

            filename

        ):

            return False

        shutil.copy2(

            filename,

            filename + ".bak"

        )

        return True

    def create_temp_file(

        self,

        suffix=".xlsx"

    ):

        fd, filename = tempfile.mkstemp(

            suffix=suffix

        )

        os.close(fd)

        return filename

    def cleanup_temp_file(

        self,

        filename

    ):

        if os.path.exists(

            filename

        ):

            os.remove(

                filename

            )

    def duplicate_workbook(

        self,

        workbook

    ):

        return copy.deepcopy(workbook)